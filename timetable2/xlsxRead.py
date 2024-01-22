#the excel reader program 8-)
#you only need 2 dictionaries theory and practical...


import openpyxl 

#use you file open function in here to insert the file...make sure the r is present with the path or else it will return an error
def read_file(filepath):
    wb = openpyxl.load_workbook(filepath)

    ws = wb.active

    #print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
    global theory
    theory = {}


    teach_all = []

    for i in range(11,(ws.max_row)): 
        v = ws.cell(row=i,column=2).value
        teach_all.append(v)
    teach = [i for i in teach_all if i is not None]    
    #print(teach)


    subs = []
    for i in range(11,292,6):
        #print(i)
        lis = []
        for j in range(i,i+6):
            #print('-',j)
            if (ws.cell(row=j,column=3).value) != None :
                lis.append(ws.cell(row=j,column=3).value)
        subs.append(lis)

    #print(subs)
    subs_load = []
    for i in range(11,292,6):
        #print(i)
        lis = []
        for j in range(i,i+6):
            #print('-',j)
            if (ws.cell(row=j,column=5).value) != None :
                lis.append(ws.cell(row=j,column=5).value)
        subs_load.append(lis)



    for i,j,k in zip(teach,subs,subs_load) :
        theory.update({i:[j,k]})
    
            

    # print("theory subjects are (teacher:subjects,load)=",theory)

    pract = []
    for i in range(11,292,6):
        #print(i)
        lis = []
        for j in range(i,i+6):
            #print('-',j)
            if (ws.cell(row=j,column=6).value) != None :
                lis.append(ws.cell(row=j,column=6).value)
        pract.append(lis)

    prac_load = []
    for i in range(11,292,6):
        #print(i)
        lis = []
        for j in range(i,i+6):
            #print('-',j)
            if (ws.cell(row=j,column=8).value) != None :
                lis.append(ws.cell(row=j,column=8).value)
        prac_load.append(lis)

    proj_load = []
    for i in range(11,292,6):
        #print(i)
        lis = []
        for j in range(i,i+6):
            #print('-',j)
            if (ws.cell(row=j,column=9).value) != None :
                lis.append(ws.cell(row=j,column=9).value)
        proj_load.append(lis)

    global practical
    practical = {}

    for i,j,k,l in zip(teach,pract,prac_load,proj_load) :
        practical.update({i:[j,k,l]})


    # print("Practical [teacher:subs,practical_load,project_load]",practical)

